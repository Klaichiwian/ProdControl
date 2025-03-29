import requests
import openpyxl
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
from datetime import datetime, timedelta
import os

# Define the API URL
API_URL = "http://167.172.68.211/api/load-order"
TOKEN = "afc1d650-024d-4615-bfbf-c01ad42ddbc8"

class ExcelAPIApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Data Fetcher")

        # Variables
        self.file_path = tk.StringVar()
        self.sheet_name = tk.StringVar()
        self.sheet_dropdown = None
        self.ws = None

        # File Selection
        tk.Label(root, text="Select Excel File:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        tk.Entry(root, textvariable=self.file_path, width=50).grid(row=0, column=1, padx=5, pady=5)
        tk.Button(root, text="Browse", command=self.browse_file).grid(row=0, column=2, padx=5, pady=5)

        # Sheet Selection
        tk.Label(root, text="Select Worksheet:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.sheet_dropdown = ttk.Combobox(root, textvariable=self.sheet_name, state="readonly")
        self.sheet_dropdown.grid(row=1, column=1, padx=5, pady=5)

        # Run Button
        tk.Button(root, text="Run", command=self.process_data).grid(row=2, column=0, columnspan=3, pady=10)

        # Output Text Box
        self.output_text = tk.Text(root, height=10, width=70)
        self.output_text.grid(row=3, column=0, columnspan=3, padx=5, pady=5)

    def browse_file(self):
        file_path = filedialog.askopenfilename(title="Select an Excel file", filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            self.file_path.set(file_path)
            self.load_sheets(file_path)

    def load_sheets(self, file_path):
        wb = openpyxl.load_workbook(file_path)
        sheets = wb.sheetnames
        self.sheet_dropdown["values"] = sheets
        if sheets:
            self.sheet_name.set(sheets[0])  # Default to first sheet

    def get_order_qty(self, pno, date):
        params = {
            "token": TOKEN,
            "dateStart": date,
            "dateEnd": date,
            "pno": pno
        }
        response = requests.get(API_URL, params=params)
        if response.status_code == 200:
            data = response.json()
            total_order_qty = sum(item['o_order_qty'] for item in data[0]['payload'])
            return total_order_qty
        return 0  # Return 0 if API fails

    def modify_dates(self):
        """ Modify dates from column F to AV, counting back into previous months if needed. """
        today = datetime.today()
        last_day = (today.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)  # Last day of month
        
        column_index = 48  # Column AV (1-based index)
        ws = self.ws

        # Fill dates backward from AV to F
        while column_index >= 6:  # Column F is index 6 (1-based)
            ws.cell(row=7, column=column_index, value=last_day.strftime('%Y-%m-%d'))
            last_day -= timedelta(days=1)  # Move to previous day
            column_index -= 1  # Move left in the sheet


    def process_data(self):
        if not self.file_path.get():
            messagebox.showerror("Error", "Please select an Excel file.")
            return

        wb = openpyxl.load_workbook(self.file_path.get())
        self.ws = wb[self.sheet_name.get()]

        # Modify date range from F to AV
        self.modify_dates()

        # Loop through column A
        for row in range(2, self.ws.max_row + 1):
            if self.ws[f"A{row}"].value is not None:  # Check if not empty
                pno = self.ws[f"B{row}"].value  # Part number

                for col in range(6, 48 + 1):  # Columns F to AV (6 to 48 in 1-based index)
                    date_cell = self.ws.cell(row=7, column=col)
                    if date_cell.value:  # Ensure the date cell is not empty
                        date_value = date_cell.value
                        if isinstance(date_value, datetime):
                            date_value = date_value.strftime('%Y-%m-%d')

                        total_order_qty = self.get_order_qty(pno, date_value)
                        self.ws.cell(row=row, column=col).value = total_order_qty

                        self.output_text.insert(tk.END, f"Written {total_order_qty} for {pno} on {date_value}\n")
                        self.output_text.see(tk.END)  # Auto-scroll

        # Save new file
        current_datetime = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        new_file_name = f"{os.path.splitext(self.file_path.get())[0]}_updated_{current_datetime}.xlsx"
        wb.save(new_file_name)

        messagebox.showinfo("Success", f"File saved as {new_file_name}")

# Run GUI
root = tk.Tk()
app = ExcelAPIApp(root)
root.mainloop()
